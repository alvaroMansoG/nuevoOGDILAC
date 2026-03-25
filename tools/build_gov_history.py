import json
import re
import csv
import io
from pathlib import Path

import openpyxl
import requests
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
TMP_DIR = DATA_DIR / "tmp"
OUTPUT_PATH = DATA_DIR / "generated" / "gov_history.json"

TMP_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "Mozilla/5.0"})

COUNTRY_ALIASES = {
    "ARG": ["Argentina"],
    "BHS": ["Bahamas(?: \\(Commonwealth of the\\))?", "Bahamas"],
    "BLZ": ["Belize"],
    "BOL": ["Bolivia(?: \\(Plurinational State(?: of)?\\))?", "Bolivia"],
    "BRA": ["Brazil(?: \\(Federative Republic of\\))?", "Brazil"],
    "BRB": ["Barbados"],
    "CHL": ["Chile"],
    "COL": ["Colombia"],
    "CRI": ["Costa Rica"],
    "DOM": ["Dominican Republic", "Dominican Rep\\.?"],
    "ECU": ["Ecuador"],
    "GTM": ["Guatemala"],
    "GUY": ["Guyana"],
    "HND": ["Honduras"],
    "HTI": ["Haiti"],
    "JAM": ["Jamaica"],
    "MEX": ["Mexico"],
    "NIC": ["Nicaragua"],
    "PAN": ["Panama"],
    "PER": ["Peru"],
    "PRY": ["Paraguay"],
    "SLV": ["El Salvador"],
    "SUR": ["Suriname(?: \\(Republic of\\))?", "Suriname"],
    "TTO": ["Trinidad and Tobago"],
    "URY": ["Uruguay(?: \\(Eastern Republic of\\))?", "Uruguay"],
    "VEN": ["Venezuela(?: \\(Bolivarian Republic of\\))?", "Venezuela"],
}

AI_TABLE_ALIASES = {
    "ARG": ["Argentina"],
    "BHS": ["Bahamas"],
    "BLZ": ["Belize"],
    "BOL": ["Bolivia"],
    "BRA": ["Brazil"],
    "BRB": ["Barbados"],
    "CHL": ["Chile"],
    "COL": ["Colombia"],
    "CRI": ["Costa Rica"],
    "DOM": ["Dominican Republic"],
    "ECU": ["Ecuador"],
    "GTM": ["Guatemala"],
    "GUY": ["Guyana"],
    "HND": ["Honduras"],
    "HTI": ["Haiti"],
    "JAM": ["Jamaica"],
    "MEX": ["Mexico"],
    "NIC": ["Nicaragua"],
    "PAN": ["Panama"],
    "PER": ["Peru"],
    "PRY": ["Paraguay"],
    "SLV": ["El Salvador"],
    "SUR": ["Suriname"],
    "TTO": ["Trinidad and Tobago"],
    "URY": ["Uruguay"],
    "VEN": ["Venezuela"],
}

ISO2_TO_ISO3 = {
    "AR": "ARG",
    "BB": "BRB",
    "BO": "BOL",
    "BR": "BRA",
    "BS": "BHS",
    "BZ": "BLZ",
    "CL": "CHL",
    "CO": "COL",
    "CR": "CRI",
    "DO": "DOM",
    "EC": "ECU",
    "GT": "GTM",
    "GY": "GUY",
    "HN": "HND",
    "HT": "HTI",
    "JM": "JAM",
    "MX": "MEX",
    "NI": "NIC",
    "PA": "PAN",
    "PE": "PER",
    "PY": "PRY",
    "SR": "SUR",
    "SV": "SLV",
    "TT": "TTO",
    "UY": "URY",
    "VE": "VEN",
}

LATAM_ISOS = list(COUNTRY_ALIASES.keys())


def normalize_text(value):
    return re.sub(r"\s+", " ", value.replace("\n", " ")).strip()


def ensure_download(url, filename):
    path = TMP_DIR / filename
    if path.exists() and path.stat().st_size > 1024:
        return path
    response = SESSION.get(url, timeout=180)
    response.raise_for_status()
    path.write_bytes(response.content)
    return path


def read_pdf_pages(path, page_range):
    reader = PdfReader(str(path))
    return normalize_text(" ".join((reader.pages[i].extract_text() or "") for i in page_range))


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    normalized = str(value).strip().replace(",", "")
    if not normalized or normalized.lower() in {"none", "n/a", "nan"}:
        return None
    try:
        return float(normalized)
    except ValueError:
        return None


def build_result():
    return {iso: {} for iso in LATAM_ISOS}


def ensure_index(result, iso, index_name):
    return result.setdefault(iso, {}).setdefault(index_name, {})


def search_country(text, aliases, patterns):
    for alias in aliases:
        for pattern in patterns:
            compiled = re.compile(pattern.format(alias=alias))
            match = compiled.search(text)
            if match:
                return match
    return None


def add_egdi_history(result):
    specs = {
        "2014": {
            "url": "https://desapublications.un.org/file/667/download",
            "filename": "egdi_2014.pdf",
            "pages": range(218, 228),
            "patterns": [r"{alias} (0\.\d+) (0\.\d+) (0\.\d+) (0\.\d+) (\d+) "],
        },
        "2016": {
            "url": "https://desapublications.un.org/file/778/download",
            "filename": "egdi_2016.pdf",
            "pages": range(166, 176),
            "patterns": [r"{alias} .*? (0\.\d+) (0\.\d+) (0\.\d+) (0\.\d+) .*? (\d+) "],
        },
        "2018": {
            "url": "https://desapublications.un.org/file/549/download",
            "filename": "egdi_2018.pdf",
            "pages": range(250, 260),
            "patterns": [r"{alias} .*? (0\.\d+) (0\.\d+) (0\.\d+) (0\.\d+) .*? (\d+) "],
        },
        "2020": {
            "url": "https://desapublications.un.org/file/781/download",
            "filename": "egdi_2020.pdf",
            "pages": range(313, 320),
            "patterns": [r"{alias} .*? (0\.\d+) (0\.\d+) (0\.\d+) (0\.\d+)"],
        },
        "2022": {
            "url": "https://desapublications.un.org/sites/default/files/publications/2022-11/Annexes%2BCover.pdf",
            "filename": "egdi_2022_annex.pdf",
            "pages": range(25, 36),
            "patterns": [r"{alias} .*? (\d+) (0\.\d+) (0\.\d+) (0\.\d+) (0\.\d+) "],
        },
    }

    for year, spec in specs.items():
        text = read_pdf_pages(ensure_download(spec["url"], spec["filename"]), spec["pages"])
        for iso, aliases in COUNTRY_ALIASES.items():
            match = search_country(text, aliases, spec["patterns"])
            if not match:
                continue

            if year in {"2014", "2016", "2018"}:
                score, osi, tii, hci, rank_world = match.groups()
            elif year == "2020":
                score, osi, tii, hci = match.groups()
                rank_world = None
            else:
                rank_world, score, osi, tii, hci = match.groups()

            ensure_index(result, iso, "egdi")[year] = {
                "score": to_number(score),
                "rankWorld": int(rank_world) if rank_world else None,
                "subindices": {
                    "osi": to_number(osi),
                    "tii": to_number(tii),
                    "hci": to_number(hci),
                },
            }


def add_gtmi_history(result):
    specs = [
        {
            "year": "2020",
            "url": "https://datacatalogfiles.worldbank.org/ddh-published/0037889/DR0065450/WBG_GovTech%20Dataset_Dec2020.xlsx",
            "filename": "gtmi_2020.xlsx",
            "sheet": "GTMI",
        },
        {
            "year": "2022",
            "url": "https://datacatalogfiles.worldbank.org/ddh-published/0037889/DR0089805/WBG_GovTech%20Dataset_Oct2022.xlsx",
            "filename": "gtmi_2022.xlsx",
            "sheet": "CG_GTMI_Groups",
        },
    ]

    for spec in specs:
        workbook = openpyxl.load_workbook(
            ensure_download(spec["url"], spec["filename"]),
            data_only=True,
            read_only=True,
        )
        worksheet = workbook[spec["sheet"]]
        rows = list(worksheet.iter_rows(values_only=True))

        entries = []
        if spec["year"] == "2020":
            for row in rows[1:]:
                if len(row) < 23 or not row[16] or to_number(row[18]) is None:
                    continue
                entries.append(
                    {
                        "iso": str(row[16]),
                        "group": str(row[17]) if row[17] else None,
                        "score": to_number(row[18]),
                        "cgsi": to_number(row[19]),
                        "psdi": to_number(row[20]),
                        "dcei": to_number(row[21]),
                        "gtei": to_number(row[22]),
                    }
                )
        else:
            headers = [header for header in rows[0]]
            header_index = {header: idx for idx, header in enumerate(headers) if header}
            for row in rows[1:]:
                iso = row[header_index["Code"]]
                score = to_number(row[header_index["GTMI"]])
                if not iso or score is None:
                    continue
                entries.append(
                    {
                        "iso": str(iso),
                        "group": str(row[header_index["Grp"]]) if row[header_index["Grp"]] else None,
                        "score": score,
                        "cgsi": to_number(row[header_index["CGSI"]]),
                        "psdi": to_number(row[header_index["PSDI"]]),
                        "dcei": to_number(row[header_index["DCEI"]]),
                        "gtei": to_number(row[header_index["GTEI"]]),
                    }
                )

        ranked = sorted((entry for entry in entries if entry["score"] is not None), key=lambda item: item["score"], reverse=True)
        rank_world = {entry["iso"]: index + 1 for index, entry in enumerate(ranked)}

        for entry in entries:
            if entry["iso"] not in result:
                continue
            ensure_index(result, entry["iso"], "gtmi")[spec["year"]] = {
                "score": entry["score"],
                "rankWorld": rank_world.get(entry["iso"]),
                "group": entry["group"],
                "subindices": {
                    "cgsi": entry["cgsi"],
                    "psdi": entry["psdi"],
                    "dcei": entry["dcei"],
                    "gtei": entry["gtei"],
                },
            }


def add_gci_history(result):
    indicator_specs = {
        "legal": "https://data360files.worldbank.org/data360-data/data/ITU_GCI/ITU_GCI_LEGL_SCRE.csv",
        "technical": "https://data360files.worldbank.org/data360-data/data/ITU_GCI/ITU_GCI_TECH_SCORE.csv",
        "organizational": "https://data360files.worldbank.org/data360-data/data/ITU_GCI/ITU_GCI_ORG_SCORE.csv",
        "cooperation": "https://data360files.worldbank.org/data360-data/data/ITU_GCI/ITU_GCI_COOP_SCORE.csv",
        "capacity": "https://data360files.worldbank.org/data360-data/data/ITU_GCI/ITU_GCI_CDS_SCORE.csv",
    }
    target_years = {"2020", "2024"}

    def is_total_row(row):
        return (
            row.get("SEX") == "_T"
            and row.get("AGE") == "_T"
            and row.get("URBANISATION") == "_T"
            and row.get("COMP_BREAKDOWN_1") == "_Z"
            and row.get("COMP_BREAKDOWN_2") == "_Z"
            and row.get("COMP_BREAKDOWN_3") == "_Z"
        )

    values_by_year = {year: {} for year in target_years}

    for sub_key, url in indicator_specs.items():
        response = SESSION.get(url, timeout=180)
        response.raise_for_status()
        reader = csv.DictReader(io.StringIO(response.text))

        for row in reader:
            year = row.get("TIME_PERIOD")
            iso = row.get("REF_AREA")
            if year not in target_years or not iso or not is_total_row(row):
                continue

            score = to_number(row.get("OBS_VALUE"))
            if score is None:
                continue

            year_values = values_by_year[year].setdefault(iso, {})
            year_values[sub_key] = score

    for year, year_values in values_by_year.items():
        ranked_entries = []
        for iso, subindices in year_values.items():
            if len(subindices) != len(indicator_specs):
                continue
            score = sum(subindices.values())
            ranked_entries.append((iso, score))

        ranked_entries.sort(key=lambda item: item[1], reverse=True)
        rank_world = {iso: index + 1 for index, (iso, _score) in enumerate(ranked_entries)}

        for iso, score in ranked_entries:
            if iso not in result:
                continue
            ensure_index(result, iso, "gci")[year] = {
                "score": score,
                "rankWorld": rank_world[iso],
                "subindices": year_values[iso],
            }


def add_ai_history(result):
    table_specs = [
        {
            "year": "2021",
            "url": "https://oxfordinsights.com/wp-content/uploads/2023/11/Government_AI_Readiness_21.pdf",
            "filename": "ai_2021.pdf",
            "pages": range(62, 72),
        },
        {
            "year": "2022",
            "url": "https://oxfordinsights.com/wp-content/uploads/2023/11/Government_AI_Readiness_2022_FV.pdf",
            "filename": "ai_2022.pdf",
            "pages": range(53, 61),
        },
        {
            "year": "2023",
            "url": "https://oxfordinsights.com/wp-content/uploads/2023/12/2023-Government-AI-Readiness-Index-2.pdf",
            "filename": "ai_2023.pdf",
            "pages": range(46, 53),
        },
        {
            "year": "2024",
            "url": "https://staging2.oxfordinsights.com/wp-content/uploads/2024/12/2024-Government-AI-Readiness-Index.pdf",
            "filename": "ai_2024.pdf",
            "pages": range(42, 50),
        },
    ]

    for spec in table_specs:
        text = read_pdf_pages(ensure_download(spec["url"], spec["filename"]), spec["pages"])
        compressed_text = text.replace(" ", "")
        patterns = [r"(\d+) {alias} (\d+\.\d+) (\d+\.\d+) (\d+\.\d+) (\d+\.\d+)"]
        for iso, aliases in AI_TABLE_ALIASES.items():
            match = search_country(text, aliases, patterns)
            if not match:
                for alias in aliases:
                    compact_alias = re.escape(alias.replace(" ", ""))
                    compact = re.search(rf"(\d+){compact_alias}(\d+\.\d+)(\d+\.\d+)(\d+\.\d+)(\d+\.\d+)", compressed_text)
                    if compact:
                        match = compact
                        break
            if not match:
                continue

            rank_world, score, government, technology, data_infrastructure = match.groups()
            ensure_index(result, iso, "ai")[spec["year"]] = {
                "score": to_number(score),
                "rankWorld": int(rank_world),
                "subindices": {
                    "government": to_number(government),
                    "technology": to_number(technology),
                    "dataInfrastructure": to_number(data_infrastructure),
                },
            }

    # 2025 is embedded in the page as JSON-like country objects.
    html = SESSION.get(
        "https://oxfordinsights.com/ai-readiness/government-ai-readiness-index-2025/",
        timeout=180,
    ).text
    object_pattern = re.compile(
        r'"name":"([^"]+)","code":"([A-Z]{2})".*?"total":([0-9.]+).*?"policy_capacity":([0-9.]+).*?"ai_infrastructure":([0-9.]+).*?"governance":([0-9.]+).*?"public_sector_adoption":([0-9.]+).*?"development_diffusion":([0-9.]+).*?"resilience":([0-9.]+)'
    )
    rows = []
    for match in object_pattern.finditer(html):
        name, iso2, total, policy, infrastructure, governance, public_sector, diffusion, resilience = match.groups()
        rows.append(
            {
                "name": name,
                "iso2": iso2,
                "score": to_number(total),
                "policyCapacity": to_number(policy),
                "aiInfrastructure": to_number(infrastructure),
                "governance": to_number(governance),
                "publicSectorAdoption": to_number(public_sector),
                "developmentDiffusion": to_number(diffusion),
                "resilience": to_number(resilience),
            }
        )

    ranked = sorted((row for row in rows if row["score"] is not None), key=lambda row: row["score"], reverse=True)
    rank_world = {row["iso2"]: index + 1 for index, row in enumerate(ranked)}
    for row in rows:
        iso = ISO2_TO_ISO3.get(row["iso2"])
        if not iso:
            continue
        ensure_index(result, iso, "ai")["2025"] = {
            "score": row["score"],
            "rankWorld": rank_world.get(row["iso2"]),
            "subindices": {
                "policyCapacity": row["policyCapacity"],
                "aiInfrastructure": row["aiInfrastructure"],
                "governance": row["governance"],
                "publicSectorAdoption": row["publicSectorAdoption"],
                "developmentDiffusion": row["developmentDiffusion"],
                "resilience": row["resilience"],
            },
        }


def add_nri_history(result):
    workbook_years = ["2021", "2022", "2023", "2024", "2025"]

    for year in workbook_years:
        workbook = openpyxl.load_workbook(
            ensure_download(
                f"https://download.networkreadinessindex.org/reports/data/{year}/nri-{year}-dataset.xlsx",
                f"nri_{year}.xlsx",
            ),
            data_only=True,
            read_only=True,
        )
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))

        header_index = None
        iso_key = None
        score_key = None
        rank_key = None
        technology_key = None
        people_key = None
        governance_key = None
        impact_key = None
        data_start = 0

        for index, row in enumerate(rows[:12]):
            normalized_headers = [str(value).strip() if value is not None else None for value in row]
            row_index = {header: idx for idx, header in enumerate(normalized_headers) if header}

            if "Code" in row_index and "NRI score" in row_index:
                header_index = row_index
                iso_key = "Code"
                score_key = "NRI score"
                rank_key = "NRI rank"
                technology_key = "Technology"
                people_key = "People"
                governance_key = "Governance"
                impact_key = "Impact"
                data_start = index + 1
                break

            if "ISO3Code" in row_index and "NRI.score" in row_index:
                header_index = row_index
                iso_key = "ISO3Code"
                score_key = "NRI.score"
                rank_key = "NRI.ranking"
                technology_key = "1.score"
                people_key = "2.score"
                governance_key = "3.score"
                impact_key = "4.score"
                data_start = index + 1
                break

        if not header_index:
            first_row = rows[0] if rows else ()
            if (
                len(first_row) > 12
                and isinstance(first_row[0], str)
                and isinstance(first_row[1], str)
                and len(str(first_row[1]).strip()) == 3
                and to_number(first_row[6]) is not None
            ):
                header_index = {}
                iso_key = 1
                score_key = 6
                rank_key = 8
                technology_key = 9
                people_key = 10
                governance_key = 11
                impact_key = 12
                data_start = 0
            else:
                workbook.close()
                raise RuntimeError(f"No se pudo detectar la cabecera del dataset NRI {year}")

        if isinstance(iso_key, int):
            required_indexes = [iso_key, score_key, rank_key, technology_key, people_key, governance_key, impact_key]
        else:
            required_indexes = [
                header_index[iso_key],
                header_index[score_key],
                header_index[rank_key],
                header_index[technology_key],
                header_index[people_key],
                header_index[governance_key],
                header_index[impact_key],
            ]

        for row in rows[data_start:]:
            if len(row) <= max(required_indexes):
                continue

            if isinstance(iso_key, int):
                iso = row[iso_key]
                score = to_number(row[score_key])
                rank_world = row[rank_key]
                technology = to_number(row[technology_key])
                people = to_number(row[people_key])
                governance = to_number(row[governance_key])
                impact = to_number(row[impact_key])
            else:
                iso = row[header_index[iso_key]]
                score = to_number(row[header_index[score_key]])
                rank_world = row[header_index[rank_key]]
                technology = to_number(row[header_index[technology_key]])
                people = to_number(row[header_index[people_key]])
                governance = to_number(row[header_index[governance_key]])
                impact = to_number(row[header_index[impact_key]])

            if not iso or str(iso) not in result:
                continue

            if score is None:
                continue

            ensure_index(result, str(iso), "nri")[year] = {
                "score": score,
                "rankWorld": int(rank_world) if rank_world is not None else None,
                "subindices": {
                    "technology": technology,
                    "people": people,
                    "governance": governance,
                    "impact": impact,
                },
            }

        workbook.close()

    # 2020 is only published as official report/country briefs PDF.
    reader = PdfReader(
        str(
            ensure_download(
                "https://download.networkreadinessindex.org/reports/data/2020/nri-2020-country-briefs.pdf",
                "nri_2020_country_briefs.pdf",
            )
        )
    )
    text_2020 = normalize_text(" ".join((page.extract_text() or "") for page in reader.pages))
    pattern_2020 = [
        r"NRI 2020 At-A-Glance: {alias} Network Readiness Index Rank: (\d+) \(out of 134\) Score:\s*([0-9.]+).*?A\. Technology pillar \d+ ([0-9.]+) C\. Governance pillar \d+ ([0-9.]+).*?B\. People pillar \d+ ([0-9.]+) D\. Impact pillar \d+ ([0-9.]+)"
    ]

    for iso, aliases in COUNTRY_ALIASES.items():
        match = search_country(text_2020, aliases, pattern_2020)
        if not match:
            continue

        rank_world, score, technology, governance, people, impact = match.groups()
        ensure_index(result, iso, "nri")["2020"] = {
            "score": to_number(score),
            "rankWorld": int(rank_world) if rank_world else None,
            "subindices": {
                "technology": to_number(technology),
                "people": to_number(people),
                "governance": to_number(governance),
                "impact": to_number(impact),
            },
        }


def prune_empty(result):
    cleaned = {}
    for iso, data in result.items():
        index_data = {}
        for index_name, series in data.items():
            if not series:
                continue
            sorted_series = {year: series[year] for year in sorted(series.keys(), key=int)}
            index_data[index_name] = sorted_series
        if index_data:
            cleaned[iso] = index_data
    return cleaned


def main():
    result = build_result()
    add_egdi_history(result)
    add_gtmi_history(result)
    add_gci_history(result)
    add_ai_history(result)
    add_nri_history(result)
    OUTPUT_PATH.write_text(json.dumps(prune_empty(result), ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
